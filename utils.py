import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import make_response
from models import AttendanceRecord, Student, BehaviorReport
from sqlalchemy import func

def export_attendance_to_excel(records, filename="attendance_report.xlsx"):
    """Export attendance records to Excel file"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الحضور"
    
    # Set RTL direction
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ['اسم الطالب', 'رقم الطالب', 'الصف', 'الفصل', 'التاريخ', 'حالة الحضور', 'ملاحظات']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.student.name)
        ws.cell(row=row, column=2, value=record.student.student_id)
        ws.cell(row=row, column=3, value=record.student.grade)
        ws.cell(row=row, column=4, value=record.student.class_name)
        ws.cell(row=row, column=5, value=record.date.strftime('%Y-%m-%d'))
        
        # Status translation
        status_map = {
            'present': 'حاضر',
            'absent': 'غائب',
            'late': 'متأخر',
            'excused': 'غياب بعذر'
        }
        ws.cell(row=row, column=6, value=status_map.get(record.status, record.status))
        ws.cell(row=row, column=7, value=record.notes or '')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

def get_monthly_attendance_stats(year, month):
    """Get attendance statistics for a specific month"""
    
    # Calculate first and last day of month
    first_day = datetime(year, month, 1).date()
    if month == 12:
        last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Get total students
    total_students = Student.query.filter_by(is_active=True).count()
    
    # Get attendance records for the month
    records = AttendanceRecord.query.filter(
        AttendanceRecord.date >= first_day,
        AttendanceRecord.date <= last_day
    ).all()
    
    # Calculate statistics
    total_days = (last_day - first_day).days + 1
    school_days = total_days  # Assuming all days are school days for simplicity
    
    present_count = len([r for r in records if r.status == 'present'])
    absent_count = len([r for r in records if r.status == 'absent'])
    late_count = len([r for r in records if r.status == 'late'])
    excused_count = len([r for r in records if r.status == 'excused'])
    
    total_possible_attendance = total_students * school_days
    attendance_rate = (present_count / total_possible_attendance * 100) if total_possible_attendance > 0 else 0
    absence_rate = (absent_count / total_possible_attendance * 100) if total_possible_attendance > 0 else 0
    
    return {
        'total_students': total_students,
        'school_days': school_days,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'attendance_rate': round(attendance_rate, 2),
        'absence_rate': round(absence_rate, 2),
        'total_records': len(records)
    }

def get_behavior_trends(student_id=None, months=6):
    """Get behavior trends for the last specified months"""
    
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=months * 30)  # Approximate
    
    query = BehaviorReport.query.filter(
        BehaviorReport.date >= start_date,
        BehaviorReport.date <= end_date
    )
    
    if student_id:
        query = query.filter(BehaviorReport.student_id == student_id)
    
    reports = query.all()
    
    # Group by month and behavior type
    monthly_data = {}
    for report in reports:
        month_key = report.date.strftime('%Y-%m')
        if month_key not in monthly_data:
            monthly_data[month_key] = {'positive': 0, 'negative': 0, 'neutral': 0}
        monthly_data[month_key][report.behavior_type] += 1
    
    # Sort by month
    sorted_months = sorted(monthly_data.keys())
    
    return {
        'months': sorted_months,
        'data': monthly_data
    }

def get_student_attendance_summary(student_id, year, month):
    """Get attendance summary for a specific student in a month"""
    
    first_day = datetime(year, month, 1).date()
    if month == 12:
        last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    records = AttendanceRecord.query.filter(
        AttendanceRecord.student_id == student_id,
        AttendanceRecord.date >= first_day,
        AttendanceRecord.date <= last_day
    ).all()
    
    summary = {
        'present': 0,
        'absent': 0,
        'late': 0,
        'excused': 0,
        'total_days': len(records)
    }
    
    for record in records:
        summary[record.status] += 1
    
    total_school_days = (last_day - first_day).days + 1
    summary['attendance_rate'] = (summary['present'] / total_school_days * 100) if total_school_days > 0 else 0
    
    return summary

def translate_status(status):
    """Translate attendance status to Arabic"""
    translations = {
        'present': 'حاضر',
        'absent': 'غائب',
        'late': 'متأخر',
        'excused': 'غياب بعذر'
    }
    return translations.get(status, status)

def translate_behavior_type(behavior_type):
    """Translate behavior type to Arabic"""
    translations = {
        'positive': 'إيجابي',
        'negative': 'سلبي',
        'neutral': 'محايد'
    }
    return translations.get(behavior_type, behavior_type)

def translate_category(category):
    """Translate behavior category to Arabic"""
    translations = {
        'discipline': 'انضباط',
        'academic': 'أكاديمي',
        'social': 'اجتماعي',
        'participation': 'مشاركة',
        'respect': 'احترام',
        'other': 'أخرى'
    }
    return translations.get(category, category)
